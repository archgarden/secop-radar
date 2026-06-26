"""Extracción de texto de documentos: PDFs digitales y escaneados."""

import os
import re
from pathlib import Path

import fitz  # PyMuPDF
import pdfplumber
import pytesseract
from PIL import Image

# Permitir configurar la ruta de Tesseract vía variable de entorno.
# En Windows local suele estar en C:\Program Files\Tesseract-OCR\tesseract.exe
TESSERACT_CMD = os.getenv("TESSERACT_CMD")
if TESSERACT_CMD:
    pytesseract.pytesseract.tesseract_cmd = TESSERACT_CMD
else:
    # Auto-detectar rutas comunes de instalación de Tesseract
    _candidatos = [
        r"C:\Program Files\Tesseract-OCR\tesseract.exe",
        r"C:\Program Files (x86)\Tesseract-OCR\tesseract.exe",
        "/usr/bin/tesseract",
        "/usr/local/bin/tesseract",
    ]
    for _c in _candidatos:
        if os.path.exists(_c):
            pytesseract.pytesseract.tesseract_cmd = _c
            break

# Configurar datos de entrenamiento (tessdata). Permite usar un directorio local del proyecto.
TESSDATA_PREFIX = os.getenv("TESSDATA_PREFIX")
if TESSDATA_PREFIX and os.path.isdir(TESSDATA_PREFIX):
    os.environ["TESSDATA_PREFIX"] = TESSDATA_PREFIX
else:
    _tessdata_local = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "tessdata")
    if os.path.isdir(_tessdata_local):
        os.environ["TESSDATA_PREFIX"] = _tessdata_local


def _tiene_texto_nativo(path: str) -> bool:
    """Determina si un PDF tiene texto seleccionable (no es solo imagen)."""
    try:
        with pdfplumber.open(path) as pdf:
            for page in pdf.pages[:3]:
                text = page.extract_text() or ""
                if text.strip():
                    return True
        return False
    except Exception:
        return False


def _extraer_texto_nativo_pdf(path: str) -> str:
    """Extrae texto de un PDF digital con pdfplumber."""
    textos = []
    try:
        with pdfplumber.open(path) as pdf:
            for page in pdf.pages:
                text = page.extract_text() or ""
                if text.strip():
                    textos.append(text)
    except Exception as exc:
        textos.append(f"[Error leyendo PDF nativo: {exc}]")
    return "\n".join(textos)


def _pdf_a_imagenes(path: str, dpi: int = 200) -> list[Image.Image]:
    """Convierte cada página de un PDF en una imagen PIL usando PyMuPDF."""
    imagenes = []
    doc = fitz.open(path)
    for page in doc:
        mat = fitz.Matrix(dpi / 72, dpi / 72)
        pix = page.get_pixmap(matrix=mat)
        img = Image.frombytes("RGB", [pix.width, pix.height], pix.samples)
        imagenes.append(img)
    doc.close()
    return imagenes


def _ocr_imagen(img: Image.Image, lang: str = "spa") -> str:
    """Aplica OCR de Tesseract a una imagen PIL."""
    return pytesseract.image_to_string(img, lang=lang)


def _extraer_texto_escaneado_pdf(path: str) -> str:
    """Renderiza páginas del PDF e aplica OCR."""
    textos = []
    try:
        imagenes = _pdf_a_imagenes(path)
        for i, img in enumerate(imagenes, 1):
            texto = _ocr_imagen(img)
            textos.append(f"--- PAGINA {i} ---\n{texto}")
    except Exception as exc:
        textos.append(f"[Error en OCR de PDF: {exc}]")
    return "\n".join(textos)


def _extraer_docx(path: str) -> str:
    try:
        from docx import Document
        doc = Document(path)
        return "\n".join(p.text for p in doc.paragraphs if p.text.strip())
    except Exception as exc:
        return f"[Error leyendo DOCX: {exc}]"


def _extraer_xlsx(path: str) -> str:
    textos = []
    try:
        from openpyxl import load_workbook
        wb = load_workbook(path, data_only=True)
        for sheet in wb.worksheets:
            for row in sheet.iter_rows(values_only=True):
                celdas = [str(c) for c in row if c is not None]
                if celdas:
                    textos.append(" ".join(celdas))
    except Exception as exc:
        textos.append(f"[Error leyendo XLSX: {exc}]")
    return "\n".join(textos)


def _extraer_txt(path: str) -> str:
    try:
        with open(path, "r", encoding="utf-8", errors="ignore") as f:
            return f.read()
    except Exception as exc:
        return f"[Error leyendo TXT: {exc}]"


def extraer_texto(path: str) -> str:
    """Extrae todo el texto posible de un archivo PDF, imagen, DOCX, XLSX o TXT."""
    ext = Path(path).suffix.lower()

    if ext == ".pdf":
        if _tiene_texto_nativo(path):
            texto = _extraer_texto_nativo_pdf(path)
            if texto.strip():
                return texto
        # Si no hay texto nativo o está vacío, usar OCR.
        return _extraer_texto_escaneado_pdf(path)

    if ext in (".png", ".jpg", ".jpeg", ".tiff", ".tif", ".bmp"):
        try:
            with Image.open(path) as img:
                return _ocr_imagen(img)
        except Exception as exc:
            return f"[Error leyendo imagen: {exc}]"

    if ext == ".docx":
        return _extraer_docx(path)

    if ext == ".xlsx":
        return _extraer_xlsx(path)

    if ext == ".txt":
        return _extraer_txt(path)

    return ""


def normalizar_texto(texto: str) -> str:
    """Limpia el texto para facilitar búsquedas por etiquetas."""
    texto = texto.lower()
    # Unificar espacios y quitar acentos comunes para matching robusto
    texto = re.sub(r"[áàäâ]", "a", texto)
    texto = re.sub(r"[éèëê]", "e", texto)
    texto = re.sub(r"[íìïî]", "i", texto)
    texto = re.sub(r"[óòöô]", "o", texto)
    texto = re.sub(r"[úùüû]", "u", texto)
    texto = re.sub(r"[ñ]", "n", texto)
    texto = re.sub(r"[^a-z0-9$.,:/\-_@%\s]", " ", texto)
    # Colapsar espacios horizontales pero preservar saltos de línea para búsquedas por línea
    texto = re.sub(r"[^\S\n]+", " ", texto)
    texto = re.sub(r"\n+", "\n", texto)
    return texto.strip()
