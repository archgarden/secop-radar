"""Análisis de pliegos y documentos SECOP."""

import json
import os
import re
import zipfile
from pathlib import Path

import pdfplumber
import pandas as pd
from docx import Document
from openpyxl import load_workbook
from sqlalchemy.orm import Session

from extraccion.ocr import extraer_texto as extraer_texto_ocr
from extraccion.requisitos_pliego import (
    extraer_requisitos_estructurados,
    resumen_requisitos_para_cliente,
)
from models import AnalisisProceso, Cliente, Documento, DocumentoProceso, Proceso

# Máximo de páginas a OCR cuando un PDF es escaneado. Los requisitos suelen estar en las primeras páginas.
OCR_MAX_PAGES = int(os.getenv("OCR_MAX_PAGES", "30"))


def extraer_texto(path: str) -> str:
    """Extrae texto de PDF, DOCX, XLSX, o ZIP (concatena todo).

    Para PDFs escaneados sin texto nativo, usa OCR con límite de páginas.
    """
    ext = Path(path).suffix.lower()

    if ext == ".pdf":
        return _extraer_pdf(path)
    if ext == ".docx":
        return _extraer_docx(path)
    if ext == ".xlsx":
        return _extraer_xlsx(path)
    if ext == ".zip":
        return _extraer_zip(path)
    if ext == ".txt":
        with open(path, "r", encoding="utf-8", errors="ignore") as f:
            return f.read()

    return ""


def _extraer_pdf(path: str) -> str:
    """Extrae texto de un PDF. Usa PyMuPDF primero por velocidad;
    si no hay texto nativo, usa OCR limitado a las primeras páginas.
    """
    try:
        import fitz  # PyMuPDF

        textos: list[str] = []
        with fitz.open(path) as doc:
            for page in doc:
                txt = page.get_text()
                if txt:
                    textos.append(txt)
        texto_nativo = "\n".join(textos)
        if texto_nativo.strip():
            return texto_nativo
    except Exception as exc:
        # Fallback a pdfplumber si PyMuPDF falla
        textos = []
        try:
            with pdfplumber.open(path) as pdf:
                for page in pdf.pages:
                    txt = page.extract_text()
                    if txt:
                        textos.append(txt)
        except Exception as exc2:
            return f"[Error leyendo PDF: {exc} / {exc2}]"
        texto_nativo = "\n".join(textos)
        if texto_nativo.strip():
            return texto_nativo

    # Si no hay texto nativo, intentar OCR (requiere Tesseract instalado).
    try:
        return _extraer_pdf_ocr(path, max_pages=OCR_MAX_PAGES)
    except Exception as exc:
        return f"[PDF sin texto nativo y OCR falló: {exc}]"


def _extraer_pdf_ocr(path: str, max_pages: int = 20) -> str:
    """Extrae texto de un PDF escaneado usando OCR, limitado a las primeras páginas."""
    import fitz  # PyMuPDF
    import pytesseract
    from PIL import Image

    textos = []
    doc = fitz.open(path)
    total = min(len(doc), max_pages)
    for i in range(total):
        page = doc[i]
        mat = fitz.Matrix(150 / 72, 150 / 72)  # 150 dpi, suficiente para OCR de texto.
        pix = page.get_pixmap(matrix=mat)
        img = Image.frombytes("RGB", [pix.width, pix.height], pix.samples)
        texto = pytesseract.image_to_string(img, lang="spa")
        if texto.strip():
            textos.append(f"--- PAGINA {i + 1} ---\n{texto}")
    doc.close()
    return "\n".join(textos)


def _extraer_docx(path: str) -> str:
    try:
        doc = Document(path)
        return "\n".join(p.text for p in doc.paragraphs if p.text.strip())
    except Exception as exc:
        return f"[Error leyendo DOCX: {exc}]"


def _extraer_xlsx(path: str) -> str:
    textos = []
    try:
        wb = load_workbook(path, data_only=True)
        for sheet in wb.worksheets:
            for row in sheet.iter_rows(values_only=True):
                celdas = [str(c) for c in row if c is not None]
                if celdas:
                    textos.append(" ".join(celdas))
    except Exception as exc:
        textos.append(f"[Error leyendo XLSX: {exc}]")
    return "\n".join(textos)


def _extraer_zip(path: str) -> str:
    textos = []
    tmp_dir = Path(path).parent / "_zip_tmp"
    tmp_dir.mkdir(exist_ok=True)
    try:
        with zipfile.ZipFile(path, "r") as z:
            for name in z.namelist():
                if name.startswith("__MACOSX") or name.startswith("."):
                    continue
                dest = tmp_dir / Path(name).name
                try:
                    z.extract(name, tmp_dir)
                    textos.append(f"--- ARCHIVO: {name} ---")
                    textos.append(extraer_texto(str(dest)))
                except Exception as exc:
                    textos.append(f"[Error extrayendo {name}: {exc}]")
    except Exception as exc:
        textos.append(f"[Error leyendo ZIP: {exc}]")
    finally:
        # Limpieza ligera
        for f in tmp_dir.iterdir():
            try:
                if f.is_file():
                    f.unlink()
                elif f.is_dir():
                    import shutil
                    shutil.rmtree(f)
            except Exception:
                pass
    return "\n".join(textos)


# Requisitos típicos en pliegos colombianos
REQUISITOS = [
    {
        "id": "rup",
        "nombre": "RUP vigente (Registro Único de Proponentes)",
        "palabras_clave": ["rup", "registro unico de proponentes", "registro único de proponentes"],
        "tipo": "legal",
    },
    {
        "id": "estados_financieros",
        "nombre": "Estados financieros con corte (año anterior)",
        "palabras_clave": ["estados financieros", "balance general", "estado de resultados", "estado de situacion financiera"],
        "tipo": "financiero",
    },
    {
        "id": "certificacion_experiencia",
        "nombre": "Certificados de experiencia en SMMLV",
        "palabras_clave": ["experiencia", "certificados de experiencia", "actas de liquidacion", "smmlv"],
        "tipo": "tecnico",
    },
    {
        "id": "parafiscales",
        "nombre": "Paz y salvo de parafiscales (SENA, ICBF, Caja)",
        "palabras_clave": ["parafiscales", "sena", "icbf", "caja de compensacion", "paz y salvo"],
        "tipo": "legal",
    },
    {
        "id": "poliza_seriedad",
        "nombre": "Póliza de seriedad de la oferta",
        "palabras_clave": ["poliza de seriedad", "seriedad de la oferta", "garantia de seriedad"],
        "tipo": "legal",
    },
    {
        "id": "propuesta_tecnica",
        "nombre": "Propuesta técnica",
        "palabras_clave": ["propuesta tecnica", "propuesta técnica", "plan de trabajo", "metodologia"],
        "tipo": "tecnico",
    },
    {
        "id": "propuesta_economica",
        "nombre": "Propuesta económica (formato de la entidad)",
        "palabras_clave": ["propuesta economica", "propuesta económica", "formato de precios", "cronograma de pagos"],
        "tipo": "economico",
    },
    {
        "id": "carta_presentacion",
        "nombre": "Carta de presentación de oferta",
        "palabras_clave": ["carta de presentacion", "carta de presentación", "oferta"],
        "tipo": "legal",
    },
    {
        "id": "capacidad_financiera",
        "nombre": "Capacidad financiera / indicadores",
        "palabras_clave": ["capacidad financiera", "liquidez", "endeudamiento", "cobertura", "patrimonio"],
        "tipo": "financiero",
    },
    {
        "id": "capacidad_residual",
        "nombre": "Capacidad residual / ejecución simultánea",
        "palabras_clave": ["capacidad residual", "ejecucion simultanea", "contratos vigentes", "obligaciones"],
        "tipo": "financiero",
    },
]


def detectar_requisitos(texto: str) -> list[dict]:
    """Detecta requisitos presentes en el texto del pliego."""
    texto_lower = texto.lower()
    resultados = []

    for req in REQUISITOS:
        encontrado = any(palabra in texto_lower for palabra in req["palabras_clave"])
        if encontrado:
            # Extraer contexto
            contextos = []
            for palabra in req["palabras_clave"]:
                for m in re.finditer(re.escape(palabra), texto_lower):
                    inicio = max(0, m.start() - 80)
                    fin = min(len(texto), m.end() + 80)
                    contextos.append(texto[inicio:fin].strip().replace("\n", " "))
                    if len(contextos) >= 2:
                        break
                if len(contextos) >= 2:
                    break

            resultados.append({
                "id": req["id"],
                "nombre": req["nombre"],
                "tipo": req["tipo"],
                "contexto": " | ".join(contextos[:2]),
            })

    return resultados


def normalizar_nombre(nombre: str) -> str:
    """Normaliza nombre de documento para comparación."""
    n = nombre.lower()
    n = re.sub(r"[^a-z0-9áéíóúñ]", " ", n)
    n = re.sub(r"\s+", " ", n).strip()
    return n


def documento_cubre_requisito(req: dict, documentos: list[Documento]) -> Documento | None:
    """Busca un documento del cliente que parezca cubrir el requisito."""
    req_nombre = normalizar_nombre(req["nombre"])
    palabras_req = set(req_nombre.split())

    for doc in documentos:
        doc_nombre = normalizar_nombre(doc.nombre)
        palabras_doc = set(doc_nombre.split())
        interseccion = palabras_req & palabras_doc
        # Coincidencia si comparten palabras clave sustanciales
        if len(interseccion) >= 2:
            return doc

        # También revisar palabras clave del requisito
        for palabra in req.get("palabras_clave", []):
            palabra_norm = normalizar_nombre(palabra)
            if palabra_norm and palabra_norm in doc_nombre:
                return doc

    return None


def analizar_pliego(proceso_id: int, cliente_id: int, db: Session) -> dict:
    """Analiza el pliego de un proceso y lo cruza con documentos del cliente."""
    proceso = db.query(Proceso).filter(Proceso.id == proceso_id).first()
    if not proceso:
        raise ValueError("Proceso no encontrado")

    cliente = db.query(Cliente).filter(Cliente.id == cliente_id).first()
    if not cliente:
        raise ValueError("Cliente no encontrado")

    documentos_cliente = db.query(Documento).filter(Documento.cliente_id == cliente_id).all()

    # Buscar un documento que parezca ser el pliego.
    # Primero intentamos con documentos descargados automáticamente del proceso.
    pliego_doc = None
    pliego_path = None
    pliego_nombre = None

    docs_proceso = (
        db.query(DocumentoProceso)
        .filter(DocumentoProceso.proceso_id == proceso_id, DocumentoProceso.es_pliego == True)
        .all()
    )
    for doc in docs_proceso:
        if doc.estado == "descargado" and doc.path and Path(doc.path).exists():
            pliego_doc = doc
            pliego_path = doc.path
            pliego_nombre = doc.nombre
            break

    # Si no hay pliego descargado, buscamos entre los documentos subidos por el cliente.
    if not pliego_path:
        for doc in documentos_cliente:
            nombre_norm = normalizar_nombre(doc.nombre)
            if "pliego" in nombre_norm or "condiciones" in nombre_norm or "terminos" in nombre_norm or "términos" in nombre_norm:
                pliego_doc = doc
                pliego_path = doc.path
                pliego_nombre = doc.nombre
                break

    if not pliego_path:
        return {
            "proceso_id": proceso_id,
            "cliente_id": cliente_id,
            "error": "No se encontró un documento identificado como pliego. Descárguelo automáticamente o súbalo manualmente.",
            "requisitos": [],
            "cumplimiento": [],
            "score_pliego": 0,
        }

    # Usar texto cacheado si existe para evitar repetir OCR costoso.
    texto_pliego = pliego_doc.texto_extraido if pliego_doc.texto_extraido else None
    if texto_pliego is None:
        texto_pliego = extraer_texto(pliego_path)
        if texto_pliego.strip() and pliego_doc:
            pliego_doc.texto_extraido = texto_pliego
            db.commit()

    if not texto_pliego.strip():
        return {
            "proceso_id": proceso_id,
            "cliente_id": cliente_id,
            "error": "No se pudo extraer texto del pliego. Verifique que el archivo no esté protegido o dañado.",
            "requisitos": [],
            "cumplimiento": [],
            "score_pliego": 0,
        }

    requisitos = detectar_requisitos(texto_pliego)

    # Extraer requisitos cuantitativos estructurados del pliego y anexos.
    todos_docs_proceso = (
        db.query(DocumentoProceso)
        .filter(DocumentoProceso.proceso_id == proceso_id)
        .all()
    )
    requisitos_estructurados = extraer_requisitos_estructurados(
        texto_pliego,
        todos_docs_proceso,
        presupuesto=proceso.presupuesto or 0,
    )
    resumen_requisitos = resumen_requisitos_para_cliente(requisitos_estructurados)

    cumplimiento = []
    cumplidos = 0
    for req in requisitos:
        doc = documento_cubre_requisito(req, documentos_cliente)
        if doc:
            cumplidos += 1
        cumplimiento.append({
            "requisito": req,
            "cumple": doc is not None,
            "documento": doc.nombre if doc else None,
            "documento_id": doc.id if doc else None,
        })

    score = round((cumplidos / len(requisitos)) * 100) if requisitos else 0

    # Guardar o actualizar en AnalisisProceso
    analisis = (
        db.query(AnalisisProceso)
        .filter(AnalisisProceso.proceso_id == proceso_id, AnalisisProceso.cliente_id == cliente_id)
        .first()
    )
    if not analisis:
        analisis = AnalisisProceso(proceso_id=proceso_id, cliente_id=cliente_id)
        db.add(analisis)

    detalle = json.loads(analisis.detalle or "{}")
    detalle["pliego"] = {
        "documento_pliego_id": pliego_doc.id if pliego_doc else None,
        "documento_pliego_nombre": pliego_nombre,
        "cantidad_requisitos": len(requisitos),
        "cantidad_cumplidos": cumplidos,
        "requisitos": requisitos,
    }
    detalle["requisitos_estructurados"] = requisitos_estructurados
    detalle["resumen_requisitos"] = resumen_requisitos

    analisis.score_pliego = score
    analisis.analisis_pliego = json.dumps({
        "cumplimiento": cumplimiento,
        "requisitos_estructurados": requisitos_estructurados,
        "resumen_requisitos": resumen_requisitos,
        "texto_pliego_preview": texto_pliego[:2000],
    })
    analisis.detalle = json.dumps(detalle)
    db.commit()
    db.refresh(analisis)

    return {
        "proceso_id": proceso_id,
        "cliente_id": cliente_id,
        "analisis_id": analisis.id,
        "documento_pliego": pliego_doc.nombre,
        "cantidad_requisitos": len(requisitos),
        "cantidad_cumplidos": cumplidos,
        "score_pliego": score,
        "requisitos": requisitos,
        "cumplimiento": cumplimiento,
        "requisitos_estructurados": requisitos_estructurados,
        "resumen_requisitos": resumen_requisitos,
    }
