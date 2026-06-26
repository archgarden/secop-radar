"""Consolida texto de todo el corpus por proceso y extrae requisitos enriquecidos.

Uso:
    cd backend
    source venv/bin/activate
    python analizar_corpus_procesos.py

Salida:
    storage/analisis_corpus/{proceso_id}.json  -> análisis por proceso
    storage/analisis_corpus_resumen.json        -> resumen global
"""

import json
import os
import re
import sys
from pathlib import Path
from typing import Any

sys.path.insert(0, str(Path(__file__).parent))

from openpyxl import load_workbook
from sqlalchemy.orm import Session

from database import SessionLocal
from extraccion.requisitos_pliego import (
    _extraer_actividad_matriz1,
    _extraer_matriz2,
    _encontrar_matriz1,
    _encontrar_matriz2,
    extraer_requisitos_estructurados,
)
from models import DocumentoProceso, Proceso

STORAGE_PATH = Path(__file__).resolve().parent.parent / "storage"
PROCESOS_DIR = STORAGE_PATH / "procesos"
OUTPUT_DIR = STORAGE_PATH / "analisis_corpus"
OUTPUT_DIR.mkdir(exist_ok=True)
RESUMEN_PATH = STORAGE_PATH / "analisis_corpus_resumen.json"

SMMLV = 1_423_500


def _normalizar(texto: str) -> str:
    t = texto.lower()
    t = t.replace("á", "a").replace("é", "e").replace("í", "i")
    t = t.replace("ó", "o").replace("ú", "u").replace("ñ", "n")
    t = re.sub(r"[^a-z0-9$.,:/\-%=\s]+", " ", t)
    t = re.sub(r"\s+", " ", t)
    return t.strip()


def _cache_path(archivo: Path) -> Path:
    return archivo.with_suffix(archivo.suffix + ".texto_extraido.txt")


def _encontrar_matriz3(documentos: list[Any]) -> str | None:
    candidatos = []
    for doc in documentos:
        nombre = (doc.nombre or doc.filename or "").lower()
        path = doc.path or ""
        if not path or not Path(path).exists():
            continue
        score = 0
        if "matriz3" in nombre or "matriz 3" in nombre:
            score += 3
        if "matriz de riesgo" in nombre or "matriz riesgos" in nombre:
            score += 2
        if "riesgos" in nombre and ("matriz3" in nombre or "matriz 3" in nombre):
            score += 1
        if any(x in nombre for x in ["experiencia", "indicador", "bienes relevantes", "bienes_relevantes"]):
            score -= 5
        if score > 0:
            candidatos.append((score, path))
    if candidatos:
        candidatos.sort(reverse=True)
        return candidatos[0][1]
    return None


def _encontrar_matriz4(documentos: list[Any]) -> str | None:
    candidatos = []
    for doc in documentos:
        nombre = (doc.nombre or doc.filename or "").lower()
        path = doc.path or ""
        if not path or not Path(path).exists():
            continue
        score = 0
        if "matriz4" in nombre or "matriz 4" in nombre:
            score += 3
        if "bienes relevantes" in nombre or "bienes_relevantes" in nombre:
            score += 2
        if "bienes" in nombre and ("matriz4" in nombre or "matriz 4" in nombre):
            score += 1
        if any(x in nombre for x in ["experiencia", "indicador", "riesgo"]):
            score -= 5
        if score > 0:
            candidatos.append((score, path))
    if candidatos:
        candidatos.sort(reverse=True)
        return candidatos[0][1]
    return None


def _extraer_matriz_excel_a_texto(path: str, max_filas: int = 50) -> dict[str, Any]:
    """Extrae las primeras filas de una matriz Excel como texto estructurado."""
    if not Path(path).exists():
        return {"path": path, "error": "No existe"}
    try:
        wb = load_workbook(path, data_only=True)
        sheet = wb.worksheets[0]
        filas = []
        for i, row in enumerate(sheet.iter_rows(values_only=True), 1):
            if i > max_filas:
                break
            celdas = [str(c) if c is not None else "" for c in row]
            if any(celdas):
                filas.append(celdas)
        return {
            "path": path,
            "filas": filas,
            "texto": "\n".join(" | ".join(f) for f in filas),
        }
    except Exception as exc:
        return {"path": path, "error": str(exc)}


def _detectar_columnas_matriz3(filas: list[list[str]]) -> list[dict[str, Any]]:
    """Detecta filas que parecen riesgos (tienen múltiples celdas no vacías)."""
    riesgos = []
    if not filas:
        return riesgos
    # La primera fila suele ser encabezado
    for fila in filas[1:]:
        celdas_no_vacias = [c for c in fila if c.strip()]
        if len(celdas_no_vacias) >= 3:
            riesgos.append({
                "descripcion": fila[0] if len(fila) > 0 else "",
                "causa": fila[1] if len(fila) > 1 else "",
                "efecto": fila[2] if len(fila) > 2 else "",
                "mitigacion": " | ".join(fila[3:]) if len(fila) > 3 else "",
            })
    return riesgos


def _detectar_columnas_matriz4(filas: list[list[str]]) -> list[dict[str, Any]]:
    """Detecta filas que parecen bienes relevantes."""
    bienes = []
    if not filas:
        return bienes
    for fila in filas[1:]:
        celdas_no_vacias = [c for c in fila if c.strip()]
        if len(celdas_no_vacias) >= 2:
            bienes.append({
                "bien": fila[0] if len(fila) > 0 else "",
                "descripcion": " | ".join(fila[1:]) if len(fila) > 1 else "",
            })
    return bienes


def _extraer_matriz3(path: str) -> dict[str, Any]:
    data = _extraer_matriz_excel_a_texto(path)
    if "error" in data:
        return data
    data["riesgos"] = _detectar_columnas_matriz3(data.get("filas", []))
    return data


def _extraer_matriz4(path: str) -> dict[str, Any]:
    data = _extraer_matriz_excel_a_texto(path)
    if "error" in data:
        return data
    data["bienes"] = _detectar_columnas_matriz4(data.get("filas", []))
    return data


def _consolidar_texto_proceso(proc_dir: Path) -> tuple[str, list[dict[str, Any]]]:
    """Concatena todo el texto extraído de un proceso."""
    fragmentos = []
    archivos = []
    for cache in sorted(proc_dir.glob("*.texto_extraido.txt")):
        original = cache.name.replace(".texto_extraido.txt", "")
        # Recuperar extensión real del original (ej: .pdf.texto_extraido.txt)
        # El cache se guarda como archivo.ext.texto_extraido.txt
        texto = cache.read_text(encoding="utf-8", errors="ignore")
        palabras = len(texto.split())
        fragmentos.append(f"\n\n=== ARCHIVO: {original} ===\n{texto}\n=== FIN {original} ===\n")
        archivos.append({
            "nombre": original,
            "palabras": palabras,
            "status": "omitido" if "OMITIDO" in texto else "extraido",
        })
    return "".join(fragmentos), archivos


def _nombre_sin_cache(nombre: str) -> str:
    """Devuelve el nombre del archivo original a partir del nombre del cache."""
    if nombre.endswith(".texto_extraido.txt"):
        return nombre[:-len(".texto_extraido.txt")]
    return nombre


def main():
    db = SessionLocal()
    try:
        procesos_dirs = [d for d in sorted(PROCESOS_DIR.iterdir()) if d.is_dir()]
        print(f"Procesos encontrados en disco: {len(procesos_dirs)}")

        resumen_global = {
            "total_procesos": len(procesos_dirs),
            "procesos_analizados": 0,
            "procesos_sin_texto": 0,
            "procesos_con_error": 0,
            "total_palabras_corpus": 0,
            "documentos_por_proceso": {},
            "requisitos_frecuentes": {},
            "advertencias_frecuentes": {},
        }

        for idx, proc_dir in enumerate(procesos_dirs, start=1):
            proceso_id = proc_dir.name
            print(f"[{idx}/{len(procesos_dirs)}] Analizando proceso {proceso_id} ...")

            # Buscar proceso en BD para presupuesto
            proceso_db = db.query(Proceso).filter(Proceso.id == int(proceso_id)).first()
            presupuesto = proceso_db.presupuesto if proceso_db else 0

            # Documentos del proceso
            docs_proceso = db.query(DocumentoProceso).filter(
                DocumentoProceso.proceso_id == int(proceso_id)
            ).all()

            texto_consolidado, archivos = _consolidar_texto_proceso(proc_dir)
            total_palabras = len(texto_consolidado.split())
            resumen_global["documentos_por_proceso"][proceso_id] = {
                "archivos": len(archivos),
                "palabras": total_palabras,
            }
            resumen_global["total_palabras_corpus"] += total_palabras

            if not texto_consolidado.strip():
                print("   → SIN TEXTO")
                resumen_global["procesos_sin_texto"] += 1
                continue

            try:
                requisitos = extraer_requisitos_estructurados(
                    texto_consolidado,
                    docs_proceso,
                    presupuesto=presupuesto,
                    smmlv=SMMLV,
                )
            except Exception as exc:
                print(f"   → ERROR extrayendo requisitos: {exc}")
                resumen_global["procesos_con_error"] += 1
                continue

            # Extraer matrices
            matriz1_path = _encontrar_matriz1(docs_proceso)
            matriz2_path = _encontrar_matriz2(docs_proceso)
            matriz3_path = _encontrar_matriz3(docs_proceso)
            matriz4_path = _encontrar_matriz4(docs_proceso)

            matrices = {
                "matriz1_experiencia": _extraer_actividad_matriz1(
                    matriz1_path,
                    requisitos.get("actividad_principal", {}).get("descripcion", ""),
                ) if matriz1_path and requisitos.get("actividad_principal") else None,
                "matriz2_indicadores": _extraer_matriz2(matriz2_path) if matriz2_path else None,
                "matriz3_riesgos": _extraer_matriz3(matriz3_path) if matriz3_path else None,
                "matriz4_bienes": _extraer_matriz4(matriz4_path) if matriz4_path else None,
            }

            # Contar frecuencia de documentos requeridos
            for doc_id in requisitos.get("documentos_requeridos", []):
                resumen_global["requisitos_frecuentes"][doc_id] = (
                    resumen_global["requisitos_frecuentes"].get(doc_id, 0) + 1
                )
            for adv in requisitos.get("advertencias", []):
                resumen_global["advertencias_frecuentes"][adv] = (
                    resumen_global["advertencias_frecuentes"].get(adv, 0) + 1
                )

            resultado = {
                "proceso_id": int(proceso_id),
                "numero_proceso": proceso_db.numero_proceso if proceso_db else None,
                "titulo": proceso_db.titulo if proceso_db else None,
                "presupuesto": presupuesto,
                "total_palabras": total_palabras,
                "archivos": archivos,
                "requisitos": requisitos,
                "matrices": matrices,
            }

            out_path = OUTPUT_DIR / f"{proceso_id}.json"
            out_path.write_text(json.dumps(resultado, indent=2, ensure_ascii=False), encoding="utf-8")

            resumen_global["procesos_analizados"] += 1
            print(f"   → OK | {total_palabras} palabras | docs: {len(archivos)}")

        # Guardar resumen global
        resumen_global["requisitos_frecuentes"] = dict(
            sorted(resumen_global["requisitos_frecuentes"].items(), key=lambda x: x[1], reverse=True)
        )
        resumen_global["advertencias_frecuentes"] = dict(
            sorted(resumen_global["advertencias_frecuentes"].items(), key=lambda x: x[1], reverse=True)
        )
        RESUMEN_PATH.write_text(json.dumps(resumen_global, indent=2, ensure_ascii=False), encoding="utf-8")

        print(f"\nAnálisis finalizado.")
        print(f"Procesos analizados: {resumen_global['procesos_analizados']}")
        print(f"Sin texto: {resumen_global['procesos_sin_texto']}")
        print(f"Con error: {resumen_global['procesos_con_error']}")
        print(f"Total palabras corpus: {resumen_global['total_palabras_corpus']:,}")
        print(f"Resumen: {RESUMEN_PATH}")

    finally:
        db.close()


if __name__ == "__main__":
    main()
